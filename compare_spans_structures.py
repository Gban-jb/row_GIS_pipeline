"""Compare Spans and Structures using per-column hash maps and Excel annotations."""

import colorsys
import importlib.util
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pandas as pd

try:
    from xlrd.biffh import XLRDError
except Exception:  # pragma: no cover - only used when xlrd is missing.
    class XLRDError(Exception):
        """Fallback xlrd error class when xlrd is unavailable at import time."""


COMMENT_AUTHOR = "ROW-GIS-Pipeline"
IGNORE_EMPTY_VALUES = True
BASE_COLOR_PALETTE: list[tuple[str, str]] = [
    ("Yellow", "FFF2CC"),
    ("Green", "D9EAD3"),
    ("Blue", "CFE2F3"),
    ("Orange", "FCE5CD"),
    ("Pink", "F4CCCC"),
    ("Teal", "D0E0E3"),
    ("Lavender", "D9D2E9"),
    ("Lime", "E2F0CB"),
    ("Peach", "FDE9D9"),
    ("Mint", "D9F2D9"),
    ("Sky", "D9EAF7"),
    ("Rose", "FADADD"),
    ("Aqua", "CCFFFF"),
    ("Gold", "FFE599"),
    ("Coral", "F4B183"),
    ("Pistachio", "D5E8D4"),
    ("PowderBlue", "DAE8FC"),
    ("Salmon", "F8CECC"),
    ("Sand", "FFF2B2"),
    ("Ice", "E1F5FE"),
    ("Seafoam", "CCF5E7"),
    ("Apricot", "FFE0B2"),
    ("Mauve", "E1BEE7"),
    ("Periwinkle", "C5CAE9"),
]


def resolve_input_file(assets_dir: Path, base_name: str) -> Path:
    """Resolve input file by preferring .xlsx and falling back to .xls."""
    preferred = assets_dir / f"{base_name}.xlsx"
    fallback = assets_dir / f"{base_name}.xls"

    if preferred.exists():
        return preferred
    if fallback.exists():
        return fallback

    raise FileNotFoundError(
        f"Could not find {base_name}.xlsx or {base_name}.xls in {assets_dir}"
    )


def ensure_dependency(package_name: str, install_hint: str) -> None:
    """Validate optional dependency availability before Excel IO steps."""
    if importlib.util.find_spec(package_name) is None:
        raise ImportError(
            f"Missing required package '{package_name}'. Install it with: {install_hint}"
        )


def read_excel_file(file_path: Path) -> pd.DataFrame:
    """Read an Excel file and handle mislabeled Excel extensions."""
    suffix = file_path.suffix.lower()

    if suffix == ".xlsx":
        ensure_dependency("openpyxl", "pip install openpyxl")
        return pd.read_excel(file_path, engine="openpyxl")

    if suffix == ".xls":
        ensure_dependency("xlrd", "pip install xlrd")
        try:
            return pd.read_excel(file_path, engine="xlrd")
        except XLRDError as exc:
            # Some files are saved as .xlsx but renamed to .xls; retry with openpyxl.
            if "xlsx file; not supported" in str(exc).lower():
                ensure_dependency("openpyxl", "pip install openpyxl")
                return pd.read_excel(file_path, engine="openpyxl")
            raise

    raise ValueError(f"Unsupported input format '{suffix}' for file: {file_path}")


def normalize_value(value: object) -> str:
    """Normalize values so comparisons are stable across blanks and mixed numeric types."""
    if pd.isna(value):
        return ""

    if isinstance(value, str):
        return value.strip()

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, ".15g")

    return str(value).strip()


def normalized_subset(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    """Return only comparison columns with NaN handled and normalized as strings."""
    if not columns:
        return pd.DataFrame(index=df.index)

    subset = df[columns].fillna("").copy()
    for column in columns:
        subset[column] = subset[column].map(normalize_value)
    return subset


def generate_pastel_hex(index: int, total: int) -> str:
    """Generate distinct pastel colors when the static palette is exhausted."""
    hue = (index % max(total, 1)) / max(total, 1)
    red, green, blue = colorsys.hls_to_rgb(hue, 0.84, 0.5)
    return f"{int(red * 255):02X}{int(green * 255):02X}{int(blue * 255):02X}"


def assign_column_colors(common_columns: list[str]) -> dict[str, dict[str, str]]:
    """Assign each common column a unique display color."""
    color_map: dict[str, dict[str, str]] = {}

    for index, column in enumerate(common_columns):
        if index < len(BASE_COLOR_PALETTE):
            color_name, hex_color = BASE_COLOR_PALETTE[index]
        else:
            color_name = f"Color_{index + 1:02d}"
            hex_color = generate_pastel_hex(index, len(common_columns))
        color_map[column] = {"name": color_name, "hex": hex_color}

    return color_map


def build_value_lookup(
    normalized_df: pd.DataFrame, columns: list[str]
) -> dict[str, dict[str, list[int]]]:
    """Build per-column hash maps: value -> list of row indices."""
    lookup: dict[str, dict[str, list[int]]] = {}

    for column in columns:
        value_to_rows: defaultdict[str, list[int]] = defaultdict(list)
        for row_index, value in enumerate(normalized_df[column].tolist()):
            if IGNORE_EMPTY_VALUES and value == "":
                continue
            value_to_rows[value].append(row_index)
        lookup[column] = dict(value_to_rows)

    return lookup


def collect_match_pairs(
    common_columns: list[str],
    structures_compare: pd.DataFrame,
    spans_compare: pd.DataFrame,
    structures_lookup: dict[str, dict[str, list[int]]],
    spans_lookup: dict[str, dict[str, list[int]]],
) -> tuple[
    dict[str, defaultdict[int, set[int]]],
    dict[str, defaultdict[int, set[int]]],
    dict[str, dict[str, int]],
]:
    """
    Collect all cross-table matched pairs for each common column using hash lookups.

    Returns:
    - structures_cell_matches: {column -> {structures_row_index -> set(spans_excel_rows)}}
    - spans_cell_matches:      {column -> {spans_row_index -> set(structures_excel_rows)}}
    - summary_stats:           per-column counts for matched rows and total pairs
    """
    structures_cell_matches: dict[str, defaultdict[int, set[int]]] = {
        column: defaultdict(set) for column in common_columns
    }
    spans_cell_matches: dict[str, defaultdict[int, set[int]]] = {
        column: defaultdict(set) for column in common_columns
    }
    summary_stats: dict[str, dict[str, int]] = {}

    for column in common_columns:
        column_pairs: set[tuple[int, int]] = set()
        spans_by_value = spans_lookup[column]
        structures_by_value = structures_lookup[column]

        # Pass 1: Structures -> Spans lookup.
        for struct_row_index, value in enumerate(structures_compare[column].tolist()):
            if IGNORE_EMPTY_VALUES and value == "":
                continue
            for span_row_index in spans_by_value.get(value, []):
                column_pairs.add((struct_row_index, span_row_index))

        # Pass 2: Spans -> Structures lookup.
        for span_row_index, value in enumerate(spans_compare[column].tolist()):
            if IGNORE_EMPTY_VALUES and value == "":
                continue
            for struct_row_index in structures_by_value.get(value, []):
                column_pairs.add((struct_row_index, span_row_index))

        matched_struct_rows: set[int] = set()
        matched_span_rows: set[int] = set()

        for struct_row_index, span_row_index in column_pairs:
            structures_cell_matches[column][struct_row_index].add(span_row_index + 2)
            spans_cell_matches[column][span_row_index].add(struct_row_index + 2)
            matched_struct_rows.add(struct_row_index + 2)
            matched_span_rows.add(span_row_index + 2)

        summary_stats[column] = {
            "Structures_Rows_Matched": len(matched_struct_rows),
            "Spans_Rows_Matched": len(matched_span_rows),
            "Total_Pairs": len(column_pairs),
        }

    return structures_cell_matches, spans_cell_matches, summary_stats


def format_row_phrase(rows: list[int]) -> str:
    """Format row labels for natural-language comment text."""
    row_labels = [f"Row {row_number}" for row_number in rows]
    if not row_labels:
        return ""
    if len(row_labels) == 1:
        return row_labels[0]
    if len(row_labels) == 2:
        return f"{row_labels[0]} and {row_labels[1]}"
    return f"{', '.join(row_labels[:-1])}, and {row_labels[-1]}"


def truncate_comment(text: str, max_len: int = 30000) -> str:
    """Limit comment text to stay within Excel comment limits."""
    if len(text) <= max_len:
        return text
    return text[: max_len - 3] + "..."


def build_row_summary_columns(
    row_count: int,
    common_columns: list[str],
    cell_match_map: dict[str, defaultdict[int, set[int]]],
    matched_row_label_prefix: str,
    detail_row_label: str,
) -> tuple[list[int], list[str], list[str]]:
    """Build Match_Count, Matched_*_Rows, and Match_Detail columns."""
    match_count_values: list[int] = []
    matched_rows_values: list[str] = []
    match_detail_values: list[str] = []

    for row_index in range(row_count):
        total_matches = 0
        matched_rows_set: set[int] = set()
        detail_parts: list[str] = []

        for column in common_columns:
            matched_rows = sorted(cell_match_map[column].get(row_index, set()))
            if not matched_rows:
                continue

            total_matches += len(matched_rows)
            matched_rows_set.update(matched_rows)
            row_csv = ",".join(str(row_number) for row_number in matched_rows)
            detail_parts.append(f"{column}->{detail_row_label}:{row_csv}")

        match_count_values.append(total_matches)
        matched_rows_values.append(
            ", ".join(
                f"{matched_row_label_prefix}{row_number}"
                for row_number in sorted(matched_rows_set)
            )
        )
        match_detail_values.append(" | ".join(detail_parts))

    return match_count_values, matched_rows_values, match_detail_values


def build_summary_sheet(
    common_columns: list[str],
    color_map: dict[str, dict[str, str]],
    summary_stats: dict[str, dict[str, int]],
) -> pd.DataFrame:
    """Create summary table with per-column match statistics and color references."""
    rows: list[dict[str, object]] = []

    for index, column in enumerate(common_columns, start=1):
        stats = summary_stats.get(
            column,
            {"Structures_Rows_Matched": 0, "Spans_Rows_Matched": 0, "Total_Pairs": 0},
        )
        color_info = color_map[column]
        rows.append(
            {
                "Column": f"{index:02d}",
                "Common_Column": column,
                "Color": f"{color_info['name']} ({color_info['hex']})",
                "Structures_Rows_Matched": stats["Structures_Rows_Matched"],
                "Spans_Rows_Matched": stats["Spans_Rows_Matched"],
                "Total_Pairs": stats["Total_Pairs"],
            }
        )

    return pd.DataFrame(
        rows,
        columns=[
            "Column",
            "Common_Column",
            "Color",
            "Structures_Rows_Matched",
            "Spans_Rows_Matched",
            "Total_Pairs",
        ],
    )


def header_index_map(worksheet) -> dict[str, int]:
    """Map worksheet header text to 1-based Excel column indices."""
    return {
        str(cell.value): column_index
        for column_index, cell in enumerate(worksheet[1], start=1)
        if cell.value is not None
    }


def add_color_legend(worksheet, color_map: dict[str, dict[str, str]]) -> None:
    """Add a right-side color legend to a worksheet."""
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    if not color_map:
        return

    start_column = worksheet.max_column + 2
    worksheet.cell(row=1, column=start_column, value="Color_Legend_Column")
    worksheet.cell(row=1, column=start_column + 1, value="Color")

    worksheet.column_dimensions[get_column_letter(start_column)].width = 28
    worksheet.column_dimensions[get_column_letter(start_column + 1)].width = 24

    for row_offset, (column_name, color_info) in enumerate(color_map.items(), start=2):
        worksheet.cell(row=row_offset, column=start_column, value=column_name)
        color_cell = worksheet.cell(
            row=row_offset,
            column=start_column + 1,
            value=f"{color_info['name']} ({color_info['hex']})",
        )
        color_cell.fill = PatternFill(
            fill_type="solid", fgColor=f"FF{color_info['hex']}"
        )


def annotate_workbook(
    workbook,
    common_columns: list[str],
    color_map: dict[str, dict[str, str]],
    structures_cell_matches: dict[str, defaultdict[int, set[int]]],
    spans_cell_matches: dict[str, defaultdict[int, set[int]]],
) -> None:
    """Apply highlights, comments, and legends to workbook sheets."""
    from openpyxl.comments import Comment
    from openpyxl.styles import PatternFill

    spans_ws = workbook["Spans"]
    structures_ws = workbook["Structures"]
    summary_ws = workbook["Summary"]

    spans_headers = header_index_map(spans_ws)
    structures_headers = header_index_map(structures_ws)
    summary_headers = header_index_map(summary_ws)

    for column in common_columns:
        color_hex = color_map[column]["hex"]
        column_fill = PatternFill(fill_type="solid", fgColor=f"FF{color_hex}")

        spans_col_index = spans_headers.get(column)
        structures_col_index = structures_headers.get(column)

        if spans_col_index is None or structures_col_index is None:
            continue

        for span_row_index, matched_struct_rows in spans_cell_matches[column].items():
            excel_row = span_row_index + 2
            matched_rows_sorted = sorted(matched_struct_rows)
            comment_text = truncate_comment(
                f"{column} matched with Structures {format_row_phrase(matched_rows_sorted)}"
            )
            cell = spans_ws.cell(row=excel_row, column=spans_col_index)
            cell.fill = column_fill
            cell.comment = Comment(comment_text, COMMENT_AUTHOR)

        for struct_row_index, matched_span_rows in structures_cell_matches[column].items():
            excel_row = struct_row_index + 2
            matched_rows_sorted = sorted(matched_span_rows)
            comment_text = truncate_comment(
                f"{column} matched with Spans {format_row_phrase(matched_rows_sorted)}"
            )
            cell = structures_ws.cell(row=excel_row, column=structures_col_index)
            cell.fill = column_fill
            cell.comment = Comment(comment_text, COMMENT_AUTHOR)

    # Shade summary color column cells with the same color assignments.
    summary_color_column = summary_headers.get("Color")
    if summary_color_column is not None:
        for row_index, column in enumerate(common_columns, start=2):
            color_hex = color_map[column]["hex"]
            summary_ws.cell(row=row_index, column=summary_color_column).fill = PatternFill(
                fill_type="solid", fgColor=f"FF{color_hex}"
            )

    add_color_legend(spans_ws, color_map)
    add_color_legend(structures_ws, color_map)


def write_output_workbook(
    preferred_output: Path,
    spans_output: pd.DataFrame,
    structures_output: pd.DataFrame,
    summary_df: pd.DataFrame,
    common_columns: list[str],
    color_map: dict[str, dict[str, str]],
    structures_cell_matches: dict[str, defaultdict[int, set[int]]],
    spans_cell_matches: dict[str, defaultdict[int, set[int]]],
) -> Path:
    """Write workbook and retry with a timestamped filename if the target is locked."""

    def _write(target_path: Path) -> None:
        with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
            spans_output.to_excel(writer, sheet_name="Spans", index=False)
            structures_output.to_excel(writer, sheet_name="Structures", index=False)
            summary_df.to_excel(writer, sheet_name="Summary", index=False)
            annotate_workbook(
                workbook=writer.book,
                common_columns=common_columns,
                color_map=color_map,
                structures_cell_matches=structures_cell_matches,
                spans_cell_matches=spans_cell_matches,
            )

    try:
        _write(preferred_output)
        return preferred_output
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback_output = preferred_output.with_name(
            f"{preferred_output.stem}_{timestamp}{preferred_output.suffix}"
        )
        _write(fallback_output)
        print(
            f"Primary output file was locked: {preferred_output.name}. "
            f"Wrote to fallback: {fallback_output.name}"
        )
        return fallback_output


def print_column_overview(spans_df: pd.DataFrame, structures_df: pd.DataFrame) -> None:
    """Print all column names at startup so available fields are visible."""
    print("\nSpans columns:")
    for index, column in enumerate(spans_df.columns, start=1):
        print(f"  {index}. {column}")

    print("\nStructures columns:")
    for index, column in enumerate(structures_df.columns, start=1):
        print(f"  {index}. {column}")


def main() -> None:
    # Resolve project paths based on this script location.
    project_root = Path(__file__).resolve().parent
    assets_dir = project_root / "assets"
    output_dir = project_root / "output"
    output_file = output_dir / "comparison_output.xlsx"

    # Locate input files (supports .xlsx and .xls names).
    spans_path = resolve_input_file(assets_dir, "Spans")
    structures_path = resolve_input_file(assets_dir, "Structures")

    # Load source workbooks into DataFrames.
    spans_df = read_excel_file(spans_path)
    structures_df = read_excel_file(structures_path)

    print(f"Loaded Spans file: {spans_path}")
    print(f"Loaded Structures file: {structures_path}")

    # Show available fields before comparison.
    print_column_overview(spans_df, structures_df)

    # Find columns shared by both tables.
    common_columns = [column for column in spans_df.columns if column in structures_df.columns]
    color_map = assign_column_colors(common_columns)

    print("\nCommon columns:")
    if common_columns:
        for index, column in enumerate(common_columns, start=1):
            color_name = color_map[column]["name"]
            color_hex = color_map[column]["hex"]
            print(f"  {index}. {column} -> {color_name} ({color_hex})")
    else:
        print("  No common columns found.")

    # Normalize values for exact string-based matching.
    spans_compare = normalized_subset(spans_df, common_columns)
    structures_compare = normalized_subset(structures_df, common_columns)

    # Build hash maps once per table: column -> value -> row_indices.
    spans_lookup = build_value_lookup(spans_compare, common_columns)
    structures_lookup = build_value_lookup(structures_compare, common_columns)

    # Collect matched cell pairs for each common column from both lookup directions.
    (
        structures_cell_matches,
        spans_cell_matches,
        summary_stats,
    ) = collect_match_pairs(
        common_columns,
        structures_compare,
        spans_compare,
        structures_lookup,
        spans_lookup,
    )

    # Build row-level summary columns for Structures.
    (
        structures_match_count,
        structures_matched_rows,
        structures_match_detail,
    ) = build_row_summary_columns(
        row_count=len(structures_df),
        common_columns=common_columns,
        cell_match_map=structures_cell_matches,
        matched_row_label_prefix="Span_Row_",
        detail_row_label="Span_Rows",
    )

    structures_output = structures_df.copy()
    structures_output["Match_Count"] = structures_match_count
    structures_output["Matched_Spans_Rows"] = structures_matched_rows
    structures_output["Match_Detail"] = structures_match_detail

    # Build row-level summary columns for Spans.
    spans_match_count, spans_matched_rows, spans_match_detail = build_row_summary_columns(
        row_count=len(spans_df),
        common_columns=common_columns,
        cell_match_map=spans_cell_matches,
        matched_row_label_prefix="Struct_Row_",
        detail_row_label="Struct_Rows",
    )

    spans_output = spans_df.copy()
    spans_output["Match_Count"] = spans_match_count
    spans_output["Matched_Struct_Rows"] = spans_matched_rows
    spans_output["Match_Detail"] = spans_match_detail

    summary_df = build_summary_sheet(common_columns, color_map, summary_stats)

    # Ensure output folder exists.
    output_dir.mkdir(parents=True, exist_ok=True)

    # openpyxl is required for writing/styling the output workbook.
    ensure_dependency("openpyxl", "pip install openpyxl")

    # Export to three sheets and apply visual annotations.
    final_output_file = write_output_workbook(
        preferred_output=output_file,
        spans_output=spans_output,
        structures_output=structures_output,
        summary_df=summary_df,
        common_columns=common_columns,
        color_map=color_map,
        structures_cell_matches=structures_cell_matches,
        spans_cell_matches=spans_cell_matches,
    )

    # Print summary statistics.
    structures_rows_with_matches = sum(1 for count in structures_match_count if count > 0)
    spans_rows_with_matches = sum(1 for count in spans_match_count if count > 0)
    total_pairs = sum(stats["Total_Pairs"] for stats in summary_stats.values())

    print("\nSummary:")
    print(f"  Common columns: {len(common_columns)}")
    print(f"  Structures rows with matches: {structures_rows_with_matches}/{len(structures_df)}")
    print(f"  Spans rows with matches: {spans_rows_with_matches}/{len(spans_df)}")
    print(f"  Total matched pairs (all columns): {total_pairs}")
    if IGNORE_EMPTY_VALUES:
        print("  Note: Empty values were ignored during matching.")
    print(f"\nOutput written to: {final_output_file}")


if __name__ == "__main__":
    main()
