"""General multi-file Excel intersection tool with compare-style output sheets."""

import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

import pandas as pd

from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

PALETTE = [
    "FFFF00",
    "92D050",
    "00B0F0",
    "FF7C80",
    "FFC000",
    "C6EFCE",
    "BDD7EE",
    "FCE4D6",
    "E2B8F0",
    "FFD966",
    "A9D18E",
    "F4B942",
    "FFDAB9",
    "DEB8FF",
    "B4E4FF",
    "D9D9D9",
]

MAX_FILES = 5
COMMENT_AUTHOR = "ROW-GIS-Pipeline"
INVALID_SHEET_CHARS = re.compile(r"[\[\]\*\?/\\:]")


def normalize_value(value: Any) -> str:
    """Apply 4-pass normalization and return comparable normalized text."""
    # Pass 1: convert value to string.
    normalized = str(value)
    # Pass 2: trim leading/trailing whitespace.
    normalized = normalized.strip()
    # Pass 3: remove trailing .0 from integer-like values.
    normalized = re.sub(r"^(-?\d+)\.0+$", r"\1", normalized)
    # Pass 4: lowercase for case-insensitive matching.
    normalized = normalized.lower()
    if normalized in {"", "nan", "none"}:
        return ""
    return normalized


def normalize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Create a normalized copy of the DataFrame while keeping raw data unchanged."""
    normalized_df = df.copy(deep=True)
    for column in normalized_df.columns:
        normalized_df[column] = normalized_df[column].map(normalize_value)
    return normalized_df


def compress_row_ranges(row_indices: List[int]) -> str:
    """Compress 0-based row indices into ranges like Row 2-5, Row 10."""
    if not row_indices:
        return "None"

    excel_rows = sorted({int(index) + 2 for index in row_indices})
    ranges: List[Tuple[int, int]] = []
    start = excel_rows[0]
    end = excel_rows[0]

    for row in excel_rows[1:]:
        if row == end + 1:
            end = row
        else:
            ranges.append((start, end))
            start = end = row
    ranges.append((start, end))

    parts = [f"Row {a}" if a == b else f"Row {a}-{b}" for a, b in ranges]
    return ", ".join(parts)


def truncate_comment(text: str, max_len: int = 30000) -> str:
    """Trim comment text to stay inside Excel comment size limits."""
    if len(text) <= max_len:
        return text
    return text[: max_len - 3] + "..."


def discover_excel_files(assets_dir: Path) -> List[Path]:
    """Find .xls/.xlsx files in assets folder and return sorted paths."""
    if not assets_dir.exists():
        print(f"ERROR: assets folder not found: {assets_dir}")
        return []

    files = [
        path
        for path in assets_dir.iterdir()
        if path.is_file() and path.suffix.lower() in {".xls", ".xlsx"}
    ]
    files.sort(key=lambda path: path.name.lower())
    return files


def load_excel_with_fallback(path: Path) -> Tuple[pd.DataFrame, str]:
    """Load Excel file using openpyxl first and xlrd as fallback."""
    last_error = None
    for engine in ["openpyxl", "xlrd"]:
        try:
            return pd.read_excel(path, engine=engine, dtype=object), engine
        except Exception as exc:  # pragma: no cover
            last_error = exc
    raise RuntimeError(f"Failed to read with openpyxl/xlrd ({last_error})")


def load_file_records(file_paths: List[Path]) -> List[Dict[str, Any]]:
    """Load input files and store raw/normalized DataFrames per file."""
    records: List[Dict[str, Any]] = []
    for path in file_paths:
        try:
            raw_df, engine = load_excel_with_fallback(path)
            norm_df = normalize_dataframe(raw_df)
            records.append(
                {
                    "name": path.name,
                    "stem": path.stem,
                    "path": path,
                    "engine": engine,
                    "raw_df": raw_df,
                    "norm_df": norm_df,
                }
            )
            print(
                f"Loaded: {path.name} | Rows: {len(raw_df)} | Columns: {len(raw_df.columns)} | Engine: {engine}"
            )
        except Exception as exc:
            print(f"ERROR: Could not load {path.name} -> {exc}")
    return records


def find_common_columns(file_records: List[Dict[str, Any]]) -> List[str]:
    """Return columns shared by all loaded files, preserving first-file order."""
    if not file_records:
        return []

    common = set(file_records[0]["raw_df"].columns)
    for record in file_records[1:]:
        common &= set(record["raw_df"].columns)
    return [column for column in file_records[0]["raw_df"].columns if column in common]


def assign_column_colors(common_columns: List[str]) -> Dict[str, str]:
    """Assign one highlight color per common column, cycling through the palette."""
    return {
        column: PALETTE[index % len(PALETTE)]
        for index, column in enumerate(common_columns)
    }


def build_lookup_maps(
    file_records: List[Dict[str, Any]], common_columns: List[str]
) -> Dict[int, Dict[str, Dict[str, List[int]]]]:
    """Build lookup[file_index][column][value] -> list of row indices."""
    lookup: Dict[int, Dict[str, Dict[str, List[int]]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(list))
    )
    for file_index, record in enumerate(file_records):
        norm_df: pd.DataFrame = record["norm_df"]
        for column in common_columns:
            for row_index, value in norm_df[column].items():
                if value != "":
                    lookup[file_index][column][value].append(int(row_index))
    return lookup


def find_verified_intersections(
    lookup: Dict[int, Dict[str, Dict[str, List[int]]]],
    common_columns: List[str],
    file_count: int,
) -> Dict[str, Dict[str, Dict[int, List[int]]]]:
    """Find values that intersect in all files with forward/reverse/cross verification."""
    intersections: Dict[str, Dict[str, Dict[int, List[int]]]] = defaultdict(dict)

    for column in common_columns:
        value_sets = [set(lookup[i][column].keys()) for i in range(file_count)]
        candidates = set.intersection(*value_sets) if value_sets and all(value_sets) else set()
        column_map: Dict[str, Dict[int, List[int]]] = {}

        for value in candidates:
            # Pass 2: forward verification.
            forward_ok = all(
                value in lookup[file_index][column] and lookup[file_index][column][value]
                for file_index in range(file_count)
            )
            if not forward_ok:
                continue

            # Pass 3: reverse verification.
            reverse_ok = True
            for file_index in range(file_count):
                current_rows = lookup[file_index][column].get(value, [])
                if not current_rows:
                    reverse_ok = False
                    break
                for other_index in range(file_count):
                    if file_index == other_index:
                        continue
                    if not lookup[other_index][column].get(value):
                        reverse_ok = False
                        break
                if not reverse_ok:
                    break

            # Pass 4: cross verification.
            if forward_ok and reverse_ok:
                column_map[value] = {
                    file_index: sorted(lookup[file_index][column][value])
                    for file_index in range(file_count)
                }

        intersections[column] = column_map

    return intersections


def build_cell_matches(
    intersections: Dict[str, Dict[str, Dict[int, List[int]]]],
    file_count: int,
) -> Dict[int, Dict[str, Dict[int, Dict[int, set]]]]:
    """Build cell-level maps: file -> column -> row -> other_file -> matched rows."""
    cell_matches: Dict[int, Dict[str, Dict[int, Dict[int, set]]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(set)))
    )

    for column, value_map in intersections.items():
        for _, per_file_rows in value_map.items():
            for file_index in range(file_count):
                for row_index in per_file_rows.get(file_index, []):
                    for other_index in range(file_count):
                        if file_index == other_index:
                            continue
                        cell_matches[file_index][column][row_index][other_index].update(
                            per_file_rows.get(other_index, [])
                        )
    return cell_matches


def make_unique_sheet_name(base_name: str, used_names: set) -> str:
    """Create a workbook-safe unique sheet name (max 31 chars)."""
    cleaned = INVALID_SHEET_CHARS.sub("_", str(base_name)).strip()
    if not cleaned:
        cleaned = "Sheet"
    candidate = cleaned[:31]
    suffix = 1
    while candidate in used_names:
        add = f"_{suffix}"
        candidate = f"{cleaned[: 31 - len(add)]}{add}"
        suffix += 1
    used_names.add(candidate)
    return candidate


def build_file_output_dataframe(
    file_index: int,
    file_records: List[Dict[str, Any]],
    common_columns: List[str],
    cell_matches: Dict[int, Dict[str, Dict[int, Dict[int, set]]]],
) -> pd.DataFrame:
    """Build compare-style output DataFrame for one file with row-level match summary."""
    raw_df = file_records[file_index]["raw_df"]
    output_df = raw_df.copy()

    match_counts: List[int] = []
    matched_rows_text: List[str] = []
    detail_text: List[str] = []

    for row_index in range(len(raw_df)):
        total_pairs = 0
        union_by_other_file: Dict[int, set] = defaultdict(set)
        column_details: List[str] = []

        for column in common_columns:
            other_map = cell_matches[file_index][column].get(row_index, {})
            if not other_map:
                continue

            for other_index, row_set in other_map.items():
                total_pairs += len(row_set)
                union_by_other_file[other_index].update(row_set)

            detail_parts = []
            for other_index in sorted(other_map.keys()):
                other_name = file_records[other_index]["name"]
                compressed = compress_row_ranges(sorted(other_map[other_index]))
                detail_parts.append(f"{other_name}: {compressed}")
            column_details.append(f"{column} -> {' | '.join(detail_parts)}")

        match_counts.append(total_pairs)

        if union_by_other_file:
            matched_parts = []
            for other_index in sorted(union_by_other_file.keys()):
                other_name = file_records[other_index]["name"]
                compressed = compress_row_ranges(sorted(union_by_other_file[other_index]))
                matched_parts.append(f"{other_name}: {compressed}")
            matched_rows_text.append(" ; ".join(matched_parts))
        else:
            matched_rows_text.append("")

        detail_text.append(" || ".join(column_details))

    output_df["Match_Count"] = match_counts
    output_df["Matched_Other_Rows"] = matched_rows_text
    output_df["Match_Detail"] = detail_text
    return output_df


def build_summary_dataframe(
    intersections: Dict[str, Dict[str, Dict[int, List[int]]]],
    file_records: List[Dict[str, Any]],
    common_columns: List[str],
    column_colors: Dict[str, str],
) -> pd.DataFrame:
    """Build summary sheet with per-column matched row counts and total cross-file pairs."""
    rows: List[Dict[str, Any]] = []
    file_count = len(file_records)

    for index, column in enumerate(common_columns, start=1):
        value_map = intersections[column]
        rows_by_file: Dict[int, set] = defaultdict(set)
        total_pairs = 0

        for _, per_file_rows in value_map.items():
            for file_index, row_indices in per_file_rows.items():
                rows_by_file[file_index].update(row_indices)
            for left in range(file_count):
                for right in range(left + 1, file_count):
                    total_pairs += len(per_file_rows.get(left, [])) * len(
                        per_file_rows.get(right, [])
                    )

        row: Dict[str, Any] = {
            "Column": f"{index:02d}",
            "Common_Column": column,
            "Color": f"#{column_colors[column]}",
            "Intersecting_Values": len(value_map),
        }
        for file_index, record in enumerate(file_records):
            row[f"{record['name']}_Rows_Matched"] = len(rows_by_file[file_index])
        row["Total_Pairs"] = total_pairs
        rows.append(row)

    ordered_columns = (
        ["Column", "Common_Column", "Color", "Intersecting_Values"]
        + [f"{record['name']}_Rows_Matched" for record in file_records]
        + ["Total_Pairs"]
    )
    return pd.DataFrame(rows, columns=ordered_columns)


def header_index_map(worksheet) -> Dict[str, int]:
    """Map worksheet header text to 1-based Excel column indices."""
    return {
        str(cell.value): column_index
        for column_index, cell in enumerate(worksheet[1], start=1)
        if cell.value is not None
    }


def add_color_legend(worksheet, column_colors: Dict[str, str]) -> None:
    """Add a right-side color legend on the worksheet."""
    if not column_colors:
        return

    start_column = worksheet.max_column + 2
    worksheet.cell(row=1, column=start_column, value="Color_Legend_Column")
    worksheet.cell(row=1, column=start_column + 1, value="Color")
    worksheet.column_dimensions[get_column_letter(start_column)].width = 32
    worksheet.column_dimensions[get_column_letter(start_column + 1)].width = 20

    for row_offset, (column_name, hex_color) in enumerate(column_colors.items(), start=2):
        worksheet.cell(row=row_offset, column=start_column, value=column_name)
        color_cell = worksheet.cell(
            row=row_offset, column=start_column + 1, value=f"#{hex_color}"
        )
        color_cell.fill = PatternFill(fill_type="solid", fgColor=f"FF{hex_color}")


def style_match_columns(worksheet) -> None:
    """Shade compare-style extra columns to match existing script behavior."""
    header_map = header_index_map(worksheet)
    highlight_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    bold_font = Font(bold=True)
    for column_name in ["Match_Count", "Matched_Other_Rows", "Match_Detail"]:
        col_idx = header_map.get(column_name)
        if col_idx is None:
            continue
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=col_idx)
            cell.fill = highlight_fill
            cell.font = bold_font


def annotate_workbook(
    workbook,
    file_records: List[Dict[str, Any]],
    file_sheet_map: Dict[int, str],
    common_columns: List[str],
    column_colors: Dict[str, str],
    cell_matches: Dict[int, Dict[str, Dict[int, Dict[int, set]]]],
) -> None:
    """Apply highlights/comments for matched cells and color legends across sheets."""
    summary_ws = workbook["Summary"]
    summary_headers = header_index_map(summary_ws)
    summary_color_index = summary_headers.get("Color")

    if summary_color_index is not None:
        for row_index, column in enumerate(common_columns, start=2):
            summary_ws.cell(row=row_index, column=summary_color_index).fill = PatternFill(
                fill_type="solid", fgColor=f"FF{column_colors[column]}"
            )

    for file_index, record in enumerate(file_records):
        worksheet = workbook[file_sheet_map[file_index]]
        headers = header_index_map(worksheet)

        style_match_columns(worksheet)
        add_color_legend(worksheet, column_colors)

        for column in common_columns:
            col_index = headers.get(column)
            if col_index is None:
                continue
            fill = PatternFill(fill_type="solid", fgColor=f"FF{column_colors[column]}")

            for row_index, other_map in cell_matches[file_index][column].items():
                excel_row = row_index + 2
                cell = worksheet.cell(row=excel_row, column=col_index)
                cell.fill = fill

                comment_lines = [f"Column : {column}", f"This cell is in: {record['name']} Row {excel_row}"]
                comment_lines.append("Matched in:")
                for other_index in sorted(other_map.keys()):
                    other_name = file_records[other_index]["name"]
                    compressed = compress_row_ranges(sorted(other_map[other_index]))
                    comment_lines.append(f"  {other_name}: {compressed}")

                cell.comment = Comment(truncate_comment("\n".join(comment_lines)), COMMENT_AUTHOR)


def write_output_workbook(
    preferred_output: Path,
    file_output_dfs: Dict[int, pd.DataFrame],
    summary_df: pd.DataFrame,
    file_records: List[Dict[str, Any]],
    common_columns: List[str],
    column_colors: Dict[str, str],
    cell_matches: Dict[int, Dict[str, Dict[int, Dict[int, set]]]],
) -> Path:
    """Write compare-style workbook and fallback to timestamped name if file is locked."""

    def _write(target_path: Path) -> Path:
        used_names = {"Summary"}
        file_sheet_map: Dict[int, str] = {}
        with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
            for file_index, record in enumerate(file_records):
                sheet_name = make_unique_sheet_name(record["stem"], used_names)
                file_sheet_map[file_index] = sheet_name
                file_output_dfs[file_index].to_excel(writer, sheet_name=sheet_name, index=False)

            summary_df.to_excel(writer, sheet_name="Summary", index=False)
            annotate_workbook(
                workbook=writer.book,
                file_records=file_records,
                file_sheet_map=file_sheet_map,
                common_columns=common_columns,
                column_colors=column_colors,
                cell_matches=cell_matches,
            )
        return target_path

    try:
        return _write(preferred_output)
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback = preferred_output.with_name(
            f"{preferred_output.stem}_{timestamp}{preferred_output.suffix}"
        )
        print(
            f"Primary output file was locked: {preferred_output.name}. "
            f"Wrote to fallback: {fallback.name}"
        )
        return _write(fallback)


def print_column_overview(file_records: List[Dict[str, Any]]) -> None:
    """Print available columns for every loaded file."""
    for record in file_records:
        print(f"\nColumns in {record['name']}:")
        for index, column in enumerate(record["raw_df"].columns, start=1):
            print(f"  {index}. {column}")


def print_common_columns(common_columns: List[str], column_colors: Dict[str, str]) -> None:
    """Print common columns and assigned colors."""
    print("\nCommon columns across all loaded files:")
    if not common_columns:
        print("  No common columns found.")
        return
    for index, column in enumerate(common_columns, start=1):
        print(f"  {index}. {column} -> #{column_colors[column]}")


def print_accuracy_report(
    file_records: List[Dict[str, Any]],
    common_columns: List[str],
    intersections: Dict[str, Dict[str, Dict[int, List[int]]]],
    file_output_dfs: Dict[int, pd.DataFrame],
    output_path: Path,
) -> None:
    """Print final accuracy report with per-column and per-file match rollups."""
    print("\n=================================================================")
    print(" ACCURACY REPORT & FINAL SUMMARY")
    print("=================================================================")
    print(f"  Files processed       : {len(file_records)}")
    for record in file_records:
        print(f"    -> {record['name']}")
    print("")
    print(f"  Common columns        : {len(common_columns)}")
    print("")
    print("  Per-column intersection results:")
    for column in common_columns:
        print(f"    {column:<20}: {len(intersections[column])} intersecting value(s)")
        for file_index, record in enumerate(file_records):
            matched_rows = set()
            for row_map in intersections[column].values():
                matched_rows.update(row_map.get(file_index, []))
            print(f"      {record['name']:<20}: {compress_row_ranges(sorted(matched_rows))}")
        print("")

    print("  Rows with matches (per file):")
    for file_index, record in enumerate(file_records):
        df = file_output_dfs[file_index]
        matched_rows = int((df["Match_Count"] > 0).sum()) if "Match_Count" in df.columns else 0
        print(f"    {record['name']:<20}: {matched_rows}/{len(df)}")

    print(f"\n  Output saved to       : {output_path}")
    print("=================================================================")
    print("  Done!")
    print("=================================================================")


def main() -> None:
    """Run compare-style multi-file intersection and create highlighted workbook output."""
    project_root = Path(__file__).resolve().parent
    assets_dir = project_root / "assets"
    output_dir = project_root / "output"
    output_path = output_dir / "intersection_output.xlsx"

    print("Scanning assets folder for Excel files...")
    file_paths = discover_excel_files(assets_dir)
    if len(file_paths) < 2:
        print("ERROR: Need at least 2 Excel files (.xls/.xlsx) in assets/.")
        return
    if len(file_paths) > MAX_FILES:
        print(f"WARNING: Found {len(file_paths)} files. Using only first {MAX_FILES}.")
        file_paths = file_paths[:MAX_FILES]

    print("\nLoading files with engine fallback...")
    file_records = load_file_records(file_paths)
    if len(file_records) < 2:
        print("ERROR: Fewer than 2 files loaded successfully. Exiting.")
        return

    print_column_overview(file_records)

    common_columns = find_common_columns(file_records)
    if not common_columns:
        print("ERROR: No common columns found across all loaded files. Exiting.")
        return

    column_colors = assign_column_colors(common_columns)
    print_common_columns(common_columns, column_colors)

    lookup = build_lookup_maps(file_records, common_columns)
    intersections = find_verified_intersections(
        lookup=lookup,
        common_columns=common_columns,
        file_count=len(file_records),
    )
    cell_matches = build_cell_matches(intersections, len(file_records))

    file_output_dfs: Dict[int, pd.DataFrame] = {}
    for file_index in range(len(file_records)):
        file_output_dfs[file_index] = build_file_output_dataframe(
            file_index=file_index,
            file_records=file_records,
            common_columns=common_columns,
            cell_matches=cell_matches,
        )

    summary_df = build_summary_dataframe(
        intersections=intersections,
        file_records=file_records,
        common_columns=common_columns,
        column_colors=column_colors,
    )

    output_dir.mkdir(parents=True, exist_ok=True)
    try:
        final_output = write_output_workbook(
            preferred_output=output_path,
            file_output_dfs=file_output_dfs,
            summary_df=summary_df,
            file_records=file_records,
            common_columns=common_columns,
            column_colors=column_colors,
            cell_matches=cell_matches,
        )
    except Exception as exc:
        print(f"ERROR: Failed to create output workbook -> {exc}")
        return

    print_accuracy_report(
        file_records=file_records,
        common_columns=common_columns,
        intersections=intersections,
        file_output_dfs=file_output_dfs,
        output_path=final_output,
    )


if __name__ == "__main__":
    main()
